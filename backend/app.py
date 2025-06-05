from flask import Flask, render_template, request, jsonify, send_file
from models.software import db, Software, Project, Release, Customer, ITHCSoftware
from flask_migrate import Migrate
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from werkzeug.utils import secure_filename
import io

def create_app(config_name='development'):
    app = Flask(__name__, 
        template_folder='../frontend/templates',
        static_folder='../frontend/static')

    # Configure database path
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'software.db')
    os.makedirs(os.path.dirname(db_path), exist_ok=True)

    if config_name == 'testing':
        app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://kratos:root/software_db'
        app.config['TESTING'] = True
    else:
        app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
    
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

    # Enable debugging
    app.debug = True

    # Database initialization
    db.init_app(app)
    migrate = Migrate(app, db)

    # Create tables only if they don't exist
    with app.app_context():
        db.create_all()

    # Configure upload folder
    UPLOAD_FOLDER = 'uploads'
    ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    @app.before_request
    def log_request_info():
        app.logger.debug('Headers: %s', request.headers)
        app.logger.debug('Body: %s', request.get_data())

    @app.after_request
    def after_request(response):
        # Add CORS headers
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE')
        
        # Only log response data for non-static files
        if not request.path.startswith('/static/'):
            try:
                app.logger.debug('Response: %s', response.get_data())
            except RuntimeError:
                app.logger.debug('Response: [Binary content]')
        
        return response

    @app.route('/')
    def index():
        return render_template('index.html')

    @app.route('/projects')
    def projects():
        return render_template('projects.html')

    @app.route('/ithc')
    def ithc():
        return render_template('ithc.html')

    @app.route('/utilities')
    def utilities():
        return render_template('utilities.html')

    @app.route('/api/software', methods=['GET'])
    def get_software():
        software_list = Software.query.all()
        return jsonify([s.to_dict() for s in software_list])

    @app.route('/api/software/<int:id>', methods=['GET'])
    def get_software_by_id(id):
        software = Software.query.get_or_404(id)
        return jsonify(software.to_dict())

    @app.route('/api/software', methods=['POST'])
    def add_software():
        try:
            data = request.json
            print("Received software data:", data)  # Debug log
            
            required_fields = ['name', 'software_type', 'latest_version']
            for field in required_fields:
                if field not in data or not data[field]:
                    print(f"Missing required field: {field}")  # Debug log
                    return jsonify({'error': f'Missing required field: {field}'}), 400

            # Check for duplicate name first
            existing = Software.query.filter_by(name=data['name']).first()
            if existing:
                print(f"Duplicate software name: {data['name']}")  # Debug log
                return jsonify({'error': 'Software with this name already exists'}), 400

            new_software = Software(
                name=data['name'],
                software_type=data['software_type'],
                latest_version=data['latest_version'],
                check_url=data.get('check_url', ''),
                last_updated=datetime.utcnow()
            )
            db.session.add(new_software)
            db.session.commit()
            print("Successfully added software:", new_software.to_dict())  # Debug log
            return jsonify(new_software.to_dict()), 201
            
        except Exception as e:
            print("Error adding software:", str(e))  # Debug log
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/software/<int:id>', methods=['PUT'])
    def update_software(id):
        software = Software.query.get_or_404(id)
        data = request.json
        software.name = data.get('name', software.name)
        software.software_type = data.get('software_type', software.software_type)
        software.latest_version = data.get('latest_version', software.latest_version)
        software.check_url = data.get('check_url', software.check_url)
        software.last_updated = datetime.utcnow()
        db.session.commit()
        return jsonify(software.to_dict())

    @app.route('/api/software/<int:id>', methods=['DELETE'])
    def delete_software(id):
        software = Software.query.get_or_404(id)
        db.session.delete(software)
        db.session.commit()
        return jsonify({'message': 'Software deleted successfully'}), 200

    @app.route('/api/software/import', methods=['POST'])
    def import_software():
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx, .xls)'}), 400

        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            workbook = load_workbook(filename=filepath)
            sheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            imported_count = 0
            for row in sheet.iter_rows(min_row=2):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if cell.value is not None}
                
                # Map Excel columns to database fields - updated mapping
                software_data = {
                    'name': row_data.get('name', row_data.get('Software')),  # Try both column names
                    'software_type': row_data.get('software_type', row_data.get('Type')),
                    'latest_version': row_data.get('latest_version', row_data.get('Latest Version')),
                    'check_url': row_data.get('check_url', row_data.get('URL'))
                }
                
                # Only process row if required fields are present
                if software_data['name'] and software_data['software_type'] and software_data['latest_version']:
                    try:
                        new_software = Software(
                            name=software_data['name'],
                            software_type=software_data['software_type'],
                            latest_version=software_data['latest_version'],
                            check_url=software_data['check_url'],
                            last_updated=datetime.utcnow()
                        )
                        db.session.add(new_software)
                        db.session.commit()
                        imported_count += 1
                    except:
                        db.session.rollback()  # Roll back on duplicate name
                
            # Clean up uploaded file
            os.remove(filepath)
            
            return jsonify({
                'message': 'Import successful',
                'imported': imported_count
            })

        except Exception as e:
            # Clean up file if it exists
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'error': f'Error processing file: {str(e)}'}), 500

    @app.route('/api/software/search', methods=['GET'])
    def search_software():
        q = request.args.get('q', '')
        software_list = Software.query.filter(
            Software.name.ilike(f'%{q}%')
        ).all()
        return jsonify([s.to_dict() for s in software_list])

    # Project routes
    @app.route('/api/projects', methods=['GET'])
    def get_projects():
        projects = Project.query.all()
        return jsonify([p.to_dict() for p in projects])

    @app.route('/api/projects/<int:id>', methods=['GET'])
    def get_project_by_id(id):
        project = Project.query.get_or_404(id)
        return jsonify(project.to_dict())

    @app.route('/api/projects', methods=['POST'])
    def add_project():
        try:
            data = request.json
            print("Received project data:", data)  # Debug log
            
            if 'name' not in data or not data['name']:
                return jsonify({'error': 'Name is required'}), 400

            # Check for duplicate name AND version combination
            existing = Project.query.filter_by(
                name=data['name'], 
                software_version=data.get('software_version')
            ).first()
            
            if existing:
                return jsonify({'error': 'Project with this name and version already exists'}), 400

            new_project = Project(
                name=data['name'],
                description=data.get('description', ''),
                software_id=data.get('software_id'),
                software_version=data.get('software_version')
            )
            
            db.session.add(new_project)
            db.session.commit()
            print("Successfully added project:", new_project.to_dict())  # Debug log
            return jsonify(new_project.to_dict()), 201
            
        except Exception as e:
            print("Error adding project:", str(e))  # Debug log
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/projects/<int:id>', methods=['PUT'])
    def update_project(id):
        try:
            project = Project.query.get_or_404(id)
            data = request.json
            
            project.name = data.get('name', project.name)
            project.description = data.get('description', project.description)
            if 'software_id' in data:
                project.software_id = data['software_id']
            if 'software_version' in data:
                project.software_version = data['software_version']
            
            db.session.commit()
            return jsonify(project.to_dict())
            
        except Exception as e:
            print("Error updating project:", str(e))
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/projects/<int:id>', methods=['DELETE'])
    def delete_project(id):
        project = Project.query.get_or_404(id)
        db.session.delete(project)
        db.session.commit()
        return jsonify({'message': 'Project deleted successfully'}), 200

    @app.route('/api/projects/search', methods=['GET'])
    def search_projects():
        q = request.args.get('q', '')
        projects = Project.query.filter(
            Project.name.ilike(f'%{q}%')
        ).all()
        return jsonify([p.to_dict() for p in projects])

    @app.route('/api/projects/import', methods=['POST'])
    def import_projects():
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx, .xls)'}), 400

        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            workbook = load_workbook(filename=filepath)
            sheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            imported_count = 0
            updated_count = 0
            skipped_count = 0
            
            for row in sheet.iter_rows(min_row=2):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if cell.value is not None}
                
                # Map Excel columns to database fields
                project_data = {
                    'name': row_data.get('Name', ''),
                    'description': row_data.get('Description', ''),
                    'software_name': row_data.get('Software Name'),
                    'software_version': row_data.get('Software Version')
                }
                
                # Only process row if name is present
                if project_data['name']:
                    try:
                        # Check for existing project with same name
                        existing = Project.query.filter_by(name=project_data['name']).first()

                        # Find associated software if specified
                        software_id = None
                        if project_data['software_name']:
                            software = Software.query.filter_by(name=project_data['software_name']).first()
                            if software:
                                software_id = software.id
                        
                        if existing:
                            # Update existing project if found
                            existing.description = project_data['description'] or existing.description
                            if software_id:
                                existing.software_id = software_id
                                existing.software_version = project_data['software_version']
                            updated_count += 1
                        else:
                            # Create new project
                            new_project = Project(
                                name=project_data['name'],
                                description=project_data['description'],
                                software_id=software_id,
                                software_version=project_data['software_version']
                            )
                            db.session.add(new_project)
                            imported_count += 1
                        
                        db.session.commit()
                    except Exception as e:
                        print(f"Error processing project {project_data['name']}: {str(e)}")
                        db.session.rollback()
                        skipped_count += 1
                
            # Clean up uploaded file
            os.remove(filepath)
            
            return jsonify({
                'message': 'Import successful',
                'imported': imported_count,
                'updated': updated_count,
                'skipped': skipped_count
            })

        except Exception as e:
            # Clean up file if it exists
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'error': f'Error processing file: {str(e)}'}), 500

    # Release routes
    @app.route('/api/projects/<int:project_id>/releases', methods=['POST'])
    def add_release(project_id):
        data = request.json
        new_release = Release(
            version=data['version'],
            notes=data.get('notes', ''),
            project_id=project_id
        )
        db.session.add(new_release)
        db.session.commit()
        return jsonify(new_release.to_dict()), 201

    @app.route('/api/releases/<int:id>', methods=['DELETE'])
    def delete_release(id):
        release = Release.query.get_or_404(id)
        db.session.delete(release)
        db.session.commit()
        return '', 204

    # Customer routes
    @app.route('/api/customers', methods=['GET'])
    def get_customers():
        customers = Customer.query.all()
        return jsonify([c.to_dict() for c in customers])

    @app.route('/api/customers', methods=['POST'])
    def add_customer():
        data = request.json
        new_customer = Customer(
            name=data['name'],
            email=data.get('email', ''),
            contact_person=data.get('contact_person', '')
        )
        db.session.add(new_customer)
        db.session.commit()
        return jsonify(new_customer.to_dict()), 201

    @app.route('/api/projects/<int:project_id>/customers/<int:customer_id>', methods=['POST'])
    def add_customer_to_project(project_id, customer_id):
        project = Project.query.get_or_404(project_id)
        customer = Customer.query.get_or_404(customer_id)
        project.customers.append(customer)
        db.session.commit()
        return jsonify(project.to_dict())

    @app.route('/api/projects/<int:project_id>/customers/<int:customer_id>', methods=['DELETE'])
    def remove_customer_from_project(project_id, customer_id):
        project = Project.query.get_or_404(project_id)
        customer = Customer.query.get_or_404(customer_id)
        project.customers.remove(customer)
        db.session.commit()
        return '', 204

    @app.route('/api/customers/import', methods=['POST'])
    def import_customers():
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx, .xls)'}), 400

        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            workbook = load_workbook(filename=filepath)
            sheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            imported_count = 0
            duplicates_count = 0
            
            for row in sheet.iter_rows(min_row=2):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if cell.value is not None}
                
                # Map Excel columns to database fields
                customer_data = {
                    'name': row_data.get('Name', row_data.get('Customer Name', row_data.get('name'))),
                    'email': row_data.get('Email', row_data.get('email', '')),
                    'contact_person': row_data.get('Contact Person', row_data.get('contact_person', ''))
                }
                
                # Only process row if name is present
                if customer_data['name']:
                    try:
                        # Check for existing customer with same name
                        existing = Customer.query.filter_by(name=customer_data['name']).first()
                        
                        if existing:
                            # Update existing customer if found
                            existing.email = customer_data['email'] or existing.email
                            existing.contact_person = customer_data['contact_person'] or existing.contact_person
                            duplicates_count += 1
                        else:
                            # Create new customer if not found
                            new_customer = Customer(
                                name=customer_data['name'],
                                email=customer_data['email'],
                                contact_person=customer_data['contact_person']
                            )
                            db.session.add(new_customer)
                            imported_count += 1
                        
                        db.session.commit()
                    except Exception as e:
                        print(f"Error processing customer {customer_data['name']}: {str(e)}")
                        db.session.rollback()
                
            # Clean up uploaded file
            os.remove(filepath)
            
            return jsonify({
                'message': 'Import successful',
                'imported': imported_count,
                'updated': duplicates_count
            })

        except Exception as e:
            # Clean up file if it exists
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'error': f'Error processing file: {str(e)}'}), 500

    # ITHC Software routes
    @app.route('/api/ithc/software', methods=['GET'])
    def get_ithc_software():
        project_id = request.args.get('project_id')
        project_version = request.args.get('project_version')
        
        query = ITHCSoftware.query
        if project_id:
            query = query.filter_by(project_id=project_id)
        if project_version:
            query = query.filter_by(project_version=project_version)
            
        ithc_list = query.all()
        return jsonify([i.to_dict() for i in ithc_list])

    @app.route('/api/ithc/software/<int:id>', methods=['GET'])
    def get_ithc_software_by_id(id):
        ithc = ITHCSoftware.query.get_or_404(id)
        return jsonify(ithc.to_dict())

    @app.route('/api/ithc/software', methods=['POST'])
    def add_ithc_software():
        try:
            data = request.json
            required_fields = ['project_id', 'software_id', 'project_version', 'current_software_version']
            for field in required_fields:
                if field not in data:
                    return jsonify({'error': f'Missing required field: {field}'}), 400

            # Check if project exists
            project = Project.query.get(data['project_id'])
            if not project:
                return jsonify({'error': 'Project not found'}), 404

            # Check if software exists
            software = Software.query.get(data['software_id'])
            if not software:
                return jsonify({'error': 'Software not found'}), 404

            # Check for duplicate entry
            existing = ITHCSoftware.query.filter_by(
                project_id=data['project_id'],
                software_id=data['software_id'],
                project_version=data['project_version']
            ).first()
            
            if existing:
                return jsonify({'error': 'Software version already exists for this project version'}), 400

            new_ithc = ITHCSoftware(
                project_id=data['project_id'],
                software_id=data['software_id'],
                project_version=data['project_version'],
                current_software_version=data['current_software_version']
            )
            db.session.add(new_ithc)
            db.session.commit()
            return jsonify(new_ithc.to_dict()), 201

        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/ithc/software/<int:id>', methods=['PUT'])
    def update_ithc_software(id):
        try:
            ithc = ITHCSoftware.query.get_or_404(id)
            data = request.json
            
            if 'current_software_version' in data:
                ithc.current_software_version = data['current_software_version']
            if 'project_version' in data:
                ithc.project_version = data['project_version']
            if 'software_id' in data:
                ithc.software_id = data['software_id']
            
            db.session.commit()
            return jsonify(ithc.to_dict())
        
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    @app.route('/api/ithc/software/<int:id>', methods=['DELETE'])
    def delete_ithc_software(id):
        ithc = ITHCSoftware.query.get_or_404(id)
        db.session.delete(ithc)
        db.session.commit()
        return '', 204

    @app.route('/api/ithc/software/search', methods=['GET'])
    def search_ithc_software():
        project_name = request.args.get('project', '')
        software_name = request.args.get('software', '')
        
        query = ITHCSoftware.query
        if project_name:
            query = query.join(Project).filter(Project.name.ilike(f'%{project_name}%'))
        if software_name:
            query = query.join(Software).filter(Software.name.ilike(f'%{software_name}%'))
            
        results = query.all()
        return jsonify([r.to_dict() for r in results])

    @app.route('/api/ithc/software/import', methods=['POST'])
    def import_ithc():
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx, .xls)'}), 400

        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            workbook = load_workbook(filename=filepath)
            sheet = workbook.active
            
            # Get headers from first row
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            
            imported_count = 0
            updated_count = 0
            skipped_count = 0
            
            for row in sheet.iter_rows(min_row=2):
                row_data = {headers[i]: str(cell.value).strip() if cell.value else '' for i, cell in enumerate(row)}
                
                try:
                    # Get project and software
                    project = Project.query.filter_by(name=row_data.get('Project Name', '')).first()
                    software = Software.query.filter_by(name=row_data.get('Software Name', '')).first()
                    
                    if not project or not software:
                        print(f"Project or software not found for row: {row_data}")
                        skipped_count += 1
                        continue

                    project_version = row_data.get('Project Version', '')
                    current_version = row_data.get('Current Version', '')

                    if not project_version or not current_version:
                        print(f"Missing version information for row: {row_data}")
                        skipped_count += 1
                        continue

                    # Check for existing entry
                    existing = ITHCSoftware.query.filter_by(
                        project_id=project.id,
                        software_id=software.id,
                        project_version=project_version
                    ).first()

                    if existing:
                        existing.current_software_version = current_version
                        updated_count += 1
                    else:
                        new_ithc = ITHCSoftware(
                            project_id=project.id,
                            software_id=software.id,
                            project_version=project_version,
                            current_software_version=current_version
                        )
                        db.session.add(new_ithc)
                        imported_count += 1

                    db.session.commit()
                except Exception as e:
                    print(f"Error processing ITHC row: {str(e)}")
                    db.session.rollback()
                    skipped_count += 1

            # Clean up uploaded file
            os.remove(filepath)
            
            return jsonify({
                'message': 'Import successful',
                'imported': imported_count,
                'updated': updated_count,
                'skipped': skipped_count
            })

        except Exception as e:
            # Clean up file if it exists
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'error': f'Error processing file: {str(e)}'}), 500

    @app.route('/api/templates/<template_type>', methods=['GET'])
    def get_template(template_type):
        try:
            wb = Workbook()
            ws = wb.active
            
            if template_type == 'software':
                ws.title = "Software Import Template"
                headers = ['Name', 'Type', 'Latest Version', 'Check URL']
                ws.append(headers)
                
            elif template_type == 'project':
                ws.title = "Project Import Template"
                headers = ['Name', 'Description', 'Software Name', 'Software Version']
                ws.append(headers)
                
            elif template_type == 'ithc':
                ws.title = "ITHC Import Template"
                headers = ['Project Name', 'Project Version', 'Software Name', 'Current Version']
                ws.append(headers)
                
            else:
                return jsonify({'error': 'Invalid template type'}), 400

            # Save to BytesIO object
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'{template_type}_template.xlsx'
            )

        except Exception as e:
            return jsonify({'error': f'Error generating template: {str(e)}'}), 500

    return app

if __name__ == '__main__':
    app = create_app()
    app.run(debug=True)
else:
    app = create_app('testing')
